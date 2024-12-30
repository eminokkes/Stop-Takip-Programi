import openpyxl

dosya = openpyxl.load_workbook("./veri.xlsx")
sayfa = dosya["veriler"]
sayfa.cell(row=1, column=2, value="ürün adı")
sayfa.cell(row=1, column=3, value="ürün alış fiyatı")
sayfa.cell(row=1, column=4, value="ürün satış fiyatı")
sayfa.cell(row=1, column=5, value="adet")
sayfa.cell(row=1, column=6, value="KDV")
sayfa.cell(row=1, column=7, value="alınan ürün sayısı")
sayfa.cell(row=1, column=8, value="satılan ürün sayısı")
sayfa.cell(row=1, column=9, value="tarih")
satir_sayisi = sayfa.max_row

for satir in range(2,satir_sayisi+1):
    ürün_alis_fiyatı=sayfa.cell(satir,3).value
    ürün_satis_fiyatı=sayfa.cell(satir,4).value
    kar = (ürün_satis_fiyatı - ürün_alis_fiyatı)
    sayfa.cell(row=1, column=10, value="kar")

for satir in range(2,satir_sayisi+1):
    ürün_alis_fiyatı=sayfa.cell(satir,3).value
    ürün_satis_fiyatı=sayfa.cell(satir,4).value
    adet=sayfa.cell(satir,5).value
    olasi_kar_toplami = ((ürün_satis_fiyatı - ürün_alis_fiyatı) * adet)
    sayfa.cell(row=1, column=11, value="olası kar toplamı")




dosya.save("./veri.xlsx")